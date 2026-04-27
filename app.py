# app.py

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, Response, abort, jsonify
from flask_talisman import Talisman
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from functools import wraps
from dotenv import load_dotenv
from odoo_manager import OdooManager
from analytics_db import AnalyticsDB
from permissions_manager import permissions_manager
from audit_logger import audit_logger
from authlib.integrations.flask_client import OAuth
from config import Config
from schemas import InventoryFilters, ExportFilters, ExportacionFilters, DashboardFilters, AnalyticsFilters
from pydantic import ValidationError
from typing import Optional, Union
from logging_config import get_logger
import os
import json
import pandas as pd
import io
import hashlib
from datetime import datetime, timedelta

load_dotenv()

logger = get_logger('app')


def _load_roles() -> dict:
    """
    Carga los roles de usuario desde roles.json (LEGACY - Ya no se usa).
    
    NOTA: Este sistema ha sido reemplazado por el módulo de permisos en base de datos.
    Se mantiene por compatibilidad pero no debe usarse en código nuevo.
    Usar permissions_manager para control de acceso.
    """
    roles_path = os.path.join(os.path.dirname(__file__), 'roles.json')
    try:
        with open(roles_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logger.warning("[CONFIG] No se pudo cargar roles.json: %s", e)
        return {'dashboard_users': [], 'admin_users': []}


ROLES = _load_roles()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

# Configuración de sesión: Expiración automática configurada
app.config['PERMANENT_SESSION_LIFETIME'] = Config.Session.LIFETIME
app.config['SESSION_COOKIE_SECURE'] = Config.Session.COOKIE_SECURE
app.config['SESSION_COOKIE_HTTPONLY'] = Config.Session.COOKIE_HTTPONLY
app.config['SESSION_COOKIE_SAMESITE'] = Config.Session.COOKIE_SAMESITE

data_manager = OdooManager()
analytics_db = AnalyticsDB()

# Rate Limiting — A01:2021 Broken Access Control (Flask-Limiter ya en requirements.txt)
limiter = Limiter(app=app, key_func=get_remote_address, default_limits=[])


# Filtro personalizado para formatear fechas
@app.template_filter('format_date')
def format_date_filter(date_str: Optional[str]) -> str:
    """
    Formatea una fecha para mostrar en formato amigable.
    Ejemplo: '2026-04-21 15:33:39' -> '21 Abr 2026 15:33'
    """
    if not date_str:
        return 'N/A'
    
    try:
        # Si viene como string con formato ISO
        if isinstance(date_str, str):
            dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        else:
            dt = date_str
        
        # Meses en español (abreviados)
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 
                 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        
        return f"{dt.day} {meses[dt.month-1]} {dt.year} {dt.strftime('%H:%M')}"
    except:
        return date_str or 'N/A'


def require_auth(f):
    """Decorador que exige sesión activa; redirige a login si no autenticado."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def require_admin(permission: Optional[str] = None):
    """
    Decorador que exige permisos de administrador.
    
    Args:
        permission: Permiso específico requerido (ej: 'manage_users', 'view_audit_log')
                   Si es None, solo verifica que sea admin_full
    
    Usage:
        @app.route('/admin/users')
        @require_admin('manage_users')
        def admin_users():
            ...
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'username' not in session:
                return redirect(url_for('login'))
            
            email = session.get('username')
            
            # Verificar si tiene el permiso específico
            if permission:
                if not permissions_manager.user_has_permission(email, permission):
                    logger.warning(f"[SECURITY] Acceso denegado a {email} - permiso requerido: {permission}")
                    abort(403)
            else:
                # Solo verificar si es admin_full
                role = permissions_manager.get_user_role(email)
                if role != 'admin_full':
                    logger.warning(f"[SECURITY] Acceso denegado a {email} - requiere admin_full, tiene: {role}")
                    abort(403)
            
            return f(*args, **kwargs)
        return decorated
    return decorator


# Configuración de OAuth2 de Google con Authlib
oauth = OAuth(app)
google = oauth.register(
    name="google",
    client_id=os.getenv('GOOGLE_CLIENT_ID'),
    client_secret=os.getenv('GOOGLE_CLIENT_SECRET'),
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)

# Configuración de seguridad OWASP: Headers con Flask-Talisman
csp = {
    'default-src': [
        "'self'",
        'https://cdn.jsdelivr.net',  # Chart.js, ECharts
        'https://cdnjs.cloudflare.com'  # html2canvas, jspdf
    ],
    'script-src': [
        "'self'",
        "'unsafe-inline'",  # Necesario para scripts inline en templates
        'https://cdn.jsdelivr.net',
        'https://cdnjs.cloudflare.com'
    ],
    'style-src': [
        "'self'",
        "'unsafe-inline'",  # Necesario para estilos inline
        'https://cdn.jsdelivr.net',
        'https://fonts.googleapis.com'
    ],
    'img-src': [
        "'self'",
        'data:',  # Para imágenes base64
        'https://*.googleusercontent.com'  # Fotos de perfil Google OAuth
    ],
    'font-src': [
        "'self'",
        'https://cdn.jsdelivr.net',
        'https://fonts.gstatic.com'
    ]
}

# Configurar Talisman solo en producción para evitar problemas en desarrollo local
if os.getenv('ENVIRONMENT', 'development') == 'production':
    Talisman(
        app,
        force_https=True,
        strict_transport_security=True,
        strict_transport_security_max_age=31536000,  # 1 año
        content_security_policy=csp,
        content_security_policy_nonce_in=['script-src'],
        feature_policy={
            'geolocation': "'none'",
            'microphone': "'none'",
            'camera': "'none'"
        }
    )

# Middleware de seguridad OWASP: Headers adicionales
@app.after_request
def set_security_headers(response: Response) -> Response:
    """
    Agrega headers de seguridad adicionales a todas las respuestas.
    
    Headers configurados:
    - X-Content-Type-Options: nosniff (previene MIME sniffing)
    - X-Frame-Options: DENY (previene clickjacking)
    - X-XSS-Protection: 1; mode=block (protección XSS legacy)
    - Referrer-Policy: strict-origin-when-cross-origin (control de referrer)
    - Permissions-Policy: desactiva APIs sensibles (geolocation, microphone, camera)
    
    Args:
        response: Objeto Response de Flask
    
    Returns:
        Response: Response modificado con headers de seguridad
    
    Note:
        Complementa la configuración de Flask-Talisman para máxima seguridad.
    """
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    return response


@app.context_processor
def inject_session_config() -> dict:
    """Inyecta constantes de sesión en todos los templates para el timer de inactividad."""
    return {
        'SESSION_TIMEOUT_MINUTES': Config.Session.TIMEOUT_MINUTES,
        'SESSION_WARNING_MINUTES': 2,
    }


@app.before_request
def log_page_visit() -> Optional[Response]:
    """
    Middleware que se ejecuta antes de cada request para verificar sesión y registrar analytics.
    
    Funcionalidad:
    - Verifica el fingerprint de sesión (previene session hijacking)
    - Verifica la expiración de sesión por inactividad
    - Actualiza el timestamp de última actividad
    - Registra visitas de página en analytics (excepto endpoints excluidos)
    - Redirige al login si la sesión expiró o fue comprometida
    
    Returns:
        Response: Redirección al login si la sesión expiró, None en caso contrario
    
    Security:
        Implementa OWASP A07:2021 - Identification and Authentication Failures
        Session fingerprinting basado en User-Agent + IP para detectar secuestro de sesión.
    """
    # Verificar fingerprint de sesión (prevención de session hijacking)
    if 'username' in session:
        # Generar fingerprint basado solo en User-Agent (la IP puede cambiar por proxies/load balancers)
        current_fingerprint = hashlib.sha256(
            f"{request.user_agent.string}".encode()
        ).hexdigest()
        
        stored_fingerprint = session.get('_security_fingerprint')
        
        if not stored_fingerprint:
            # Primera vez, almacenar fingerprint
            session['_security_fingerprint'] = current_fingerprint
        elif current_fingerprint != stored_fingerprint:
            # Fingerprint no coincide - posible session hijacking
            logger.warning(
                "[SECURITY] Posible session hijacking | user=%s | ip=%s | path=%s",
                session.get('username'), request.remote_addr, request.path
            )
            session.clear()
            flash('Sesión inválida detectada. Por favor, inicia sesión nuevamente.', 'danger')
            return redirect(url_for('login'))
    
    # Verificar expiración de sesión por inactividad
    if 'username' in session:
        last_activity = session.get('last_activity')
        if last_activity:
            try:
                last_activity_time = datetime.fromisoformat(last_activity)
                inactive_time = datetime.now() - last_activity_time
                
                # Si han pasado más del tiempo configurado de inactividad, cerrar sesión
                if inactive_time > Config.Session.LIFETIME:
                    logger.info(
                        "[SECURITY] Sesión expirada por inactividad | user=%s | inactivo=%.1f min",
                        session.get('username'), inactive_time.total_seconds() / 60
                    )
                    session.clear()
                    flash('Tu sesión ha expirado por inactividad. Por favor, inicia sesión nuevamente.', 'warning')
                    return redirect(url_for('login'))
            except (ValueError, TypeError):
                pass
        
        # Actualizar última actividad
        session['last_activity'] = datetime.now().isoformat()
        
        # Registrar visitas (excluyendo endpoints específicos)
        if request.endpoint not in ['static', None]:
            # Excluir usuario administrador de analytics (temporalmente desactivado para pruebas)
            excluded_users = []  # jonathan.cerda@agrovetmarket.com - desactivado para pruebas
            if session.get('username').lower() in [email.lower() for email in excluded_users]:
                return  # No registrar visitas del administrador
            
            # Excluir endpoints que no son páginas reales
            excluded_endpoints = ['static', 'export_excel', 'export_excel_exportacion']
            if request.endpoint not in excluded_endpoints:
                analytics_db.log_visit(
                    user_email=session.get('username'),
                    user_name=session.get('username').split('@')[0],  # Nombre simple del email
                    page_url=request.path,
                    page_title=request.endpoint,
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent'),
                    referrer=request.referrer,
                    method=request.method
                )

@app.route('/login')
def login() -> str:
    """
    Renderiza la página de inicio de sesión.
    
    Returns:
        str: Template HTML de la página de login
    """
    return render_template('login.html')

@app.route('/google-oauth')
@limiter.limit('5 per minute')
def google_oauth() -> Response:
    """
    Inicia el flujo de autenticación OAuth2 con Google.
    
    Redirige al usuario a la página de autenticación de Google para
    obtener autorización y acceso a su información de perfil.
    
    Returns:
        Response: Redirección a Google OAuth2 authorization endpoint
    """
    redirect_uri = url_for('authorize', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route('/authorize')
def authorize() -> Response:
    """
    Callback de OAuth2 que procesa la autorización de Google.
    
    Funcionalidad:
    - Obtiene el token de acceso de Google
    - Extrae información del usuario (email, nombre, foto)
    - Verifica si el usuario está registrado y activo en la base de datos
    - Registra el último login del usuario
    - Crea la sesión del usuario con su rol
    - Redirige según permisos del usuario (dashboard o inventario)
    
    Returns:
        Response: Redirección al dashboard/inventario si exitoso, al login si falla
        
    Raises:
        Exception: Captura errores de autenticación OAuth2
    
    Security:
        - Solo usuarios registrados en inventario_user_permissions pueden acceder
        - Usuarios inactivos son rechazados
        - Todos los intentos de acceso son logueados
    """
    try:
        token = google.authorize_access_token()
        user_info = token.get('userinfo')
        
        if not user_info:
            flash('No se pudo obtener información del usuario.', 'danger')
            return redirect(url_for('login'))
        
        email = user_info.get('email')
        name = user_info.get('name')
        picture = user_info.get('picture')
        
        # Verificar si el usuario está registrado en el sistema de permisos
        user_details = permissions_manager.get_user_details(email)
        
        if not user_details or not user_details.get('is_active'):
            logger.warning(
                "[SECURITY] Intento de acceso no autorizado | email=%s | ip=%s | registered=%s",
                email, request.remote_addr, bool(user_details)
            )
            flash('Acceso denegado. Tu cuenta no está registrada o está inactiva. Contacta al administrador.', 'danger')
            return redirect(url_for('login'))
        
        # Registrar último login del usuario
        permissions_manager.update_last_login(email)
        
        # Guardar la información del usuario en la sesión
        session.permanent = True  # Activar expiración automática de 15 minutos
        session['username'] = email
        session['user_name'] = name
        session['user_picture'] = picture
        session['user_info'] = user_info
        session['user_role'] = user_details.get('role')
        session['last_activity'] = datetime.now().isoformat()
        
        flash(f'¡Bienvenido/a {name}!', 'success')
        
        # Redirigir según permisos del usuario
        if permissions_manager.user_has_permission(email, 'view_dashboard'):
            return redirect(url_for('dashboard'))
        elif permissions_manager.user_has_permission(email, 'view_inventory'):
            return redirect(url_for('inventory'))
        else:
            flash('No tienes permisos asignados. Contacta al administrador.', 'warning')
            return redirect(url_for('login'))
        
    except Exception as e:
        logger.error(f"Error en autenticación OAuth2: {e}", exc_info=True)
        flash(f'Error durante la autenticación: {str(e)}', 'danger')
        return redirect(url_for('login'))

@app.route('/logout')
def logout() -> Response:
    """
    Cierra la sesión del usuario actual.
    
    Limpia toda la información de sesión almacenada y redirige
    a la página de login.
    
    Returns:
        Response: Redirección a la página de login
    """
    session.clear()
    flash('Has cerrado la sesión.', 'info')
    return redirect(url_for('login'))


@app.route('/api/keep-alive', methods=['POST'])
def keep_alive():
    """
    Renueva la sesión activa para evitar cierre por inactividad.

    Llamado por el timer de inactividad del frontend cuando el usuario
    confirma que desea continuar. El middleware before_request ya actualiza
    last_activity en cada request, por lo que basta con retornar OK.

    Returns:
        JSON {ok: True} si la sesión sigue activa, 401 si ya expiró.
    """
    if 'username' not in session:
        return {'ok': False, 'expired': True}, 401
    return {'ok': True}


@app.route('/export/excel/exportacion')
@require_auth
@limiter.limit('10 per hour')
def export_excel_exportacion() -> Union[Response, str]:
    """
    Genera y descarga un archivo Excel con el inventario de la ubicación de exportación.
    
    Consulta productos de la ubicación 'ALMC/Stock/PCP/Exportacion' con filtros
    opcionales y genera un archivo Excel formateado.
    
    Query Parameters:
        search_term (str, optional): Término de búsqueda para filtrar productos
        category_id (int, optional): ID del grupo de artículos para filtrar
        line_id (int, optional): ID de la línea comercial para filtrar
    
    Returns:
        Response: Archivo Excel para descarga o redirigir si no hay datos
        
    Requires:
        Sesión de usuario activa (username in session)
    """
    try:
        f = ExportacionFilters.model_validate({
            'search_term': request.args.get('search_term'),
            'category_id': request.args.get('category_id'),
            'line_id': request.args.get('line_id'),
        })
    except ValidationError:
        abort(400)
    inventory_data = data_manager.get_export_inventory(
        search_term=f.search_term,
        category_id=f.category_id,
        line_id=f.line_id,
    )

    if not inventory_data:
        flash('No hay datos para exportar.', 'warning')
        return redirect(url_for('inventory_export'))
        
    df = pd.DataFrame([{k: v for k, v in item.items() if k not in ['product_id', 'grupo_articulo_id', 'linea_comercial_id']} for item in inventory_data])

    # Formatear columnas numéricas y de fecha
    if 'cantidad_disponible' in df.columns:
        df['cantidad_disponible'] = pd.to_numeric(df['cantidad_disponible'].str.replace(',', ''), errors='coerce')
    if 'fecha_expira' in df.columns:
        df['fecha_expira'] = pd.to_datetime(df['fecha_expira'], format='%d-%m-%Y', errors='coerce')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Inventario Exportacion')
        ws = writer.sheets['Inventario Exportacion']
        # Ajustar ancho de columnas automáticamente
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        # Formato para cantidades y fechas
        from openpyxl.styles import numbers
        if 'cantidad_disponible' in df.columns:
            col_idx = df.columns.get_loc('cantidad_disponible') + 1
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for excel_cell in cell:
                    excel_cell.number_format = Config.Export.NUMBER_FORMAT
        if 'fecha_expira' in df.columns:
            col_idx = df.columns.get_loc('fecha_expira') + 1
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for excel_cell in cell:
                    excel_cell.number_format = 'DD-MM-YYYY'
    output.seek(0)

    filename = f"inventario_exportacion_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

@app.route('/export/excel')
@require_auth
@limiter.limit('10 per hour')
def export_excel() -> Union[Response, str]:
    """
    Genera y descarga un archivo Excel con el inventario general.
    
    Consulta el inventario completo con múltiples filtros opcionales y
    genera un archivo Excel formateado con cantidades y fechas.
    
    Query Parameters:
        search_term (str, optional): Término de búsqueda para filtrar productos
        product_id (int, optional): ID específico de producto
        category_id (int, optional): ID del grupo de artículos
        line_id (int, optional): ID de la línea comercial
        location_id (int, optional): ID de la ubicación de almacén
        exp_status (str, optional): Estado de expiración (vence_pronto, advertencia, ok, largo_plazo)
    
    Returns:
        Response: Archivo Excel para descarga o redirección si no hay datos
        
    Requires:
        Sesión de usuario activa (username in session)
    """
    try:
        f = ExportFilters.model_validate({
            'search_term': request.args.get('search_term'),
            'product_id': request.args.get('product_id'),
            'category_id': request.args.get('category_id'),
            'line_id': request.args.get('line_id'),
            'location_id': request.args.get('location_id'),
            'exp_status': request.args.get('exp_status'),
        })
    except ValidationError:
        abort(400)
    inventory_data = data_manager.get_stock_inventory(
        search_term=f.search_term,
        product_id=f.product_id,
        category_id=f.category_id,
        line_id=f.line_id,
        location_id=f.location_id,
    )

    exp_status = f.exp_status
    if exp_status and inventory_data:
        if exp_status == 'vence_pronto':
            inventory_data = [item for item in inventory_data if item['meses_expira'] is not None and 0 <= item['meses_expira'] <= 3]
        elif exp_status == 'advertencia':
            inventory_data = [item for item in inventory_data if item['meses_expira'] is not None and 4 <= item['meses_expira'] <= 7]
        elif exp_status == 'ok':
            inventory_data = [item for item in inventory_data if item['meses_expira'] is not None and 8 <= item['meses_expira'] <= 12]
        elif exp_status == 'largo_plazo':
            inventory_data = [item for item in inventory_data if item['meses_expira'] is not None and item['meses_expira'] > 12]

    if not inventory_data:
        flash('No hay datos para exportar.', 'warning')
        return redirect(url_for('inventory'))
    
    export_df_data = [{k: v for k, v in item.items() if k not in ['product_id', 'grupo_articulo_id', 'linea_comercial_id']} for item in inventory_data]
    df = pd.DataFrame(export_df_data)

    # Formatear columnas numéricas y de fecha
    if 'cantidad_disponible' in df.columns:
        df['cantidad_disponible'] = pd.to_numeric(df['cantidad_disponible'].astype(str).str.replace(',', ''), errors='coerce')
    if 'fecha_expira' in df.columns:
        df['fecha_expira'] = pd.to_datetime(df['fecha_expira'], format='%d-%m-%Y', errors='coerce')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Inventario')
        ws = writer.sheets['Inventario']
        # Ajustar ancho de columnas automáticamente
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        # Formato para cantidades y fechas
        from openpyxl.styles import numbers
        if 'cantidad_disponible' in df.columns:
            col_idx = df.columns.get_loc('cantidad_disponible') + 1
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for excel_cell in cell:
                    excel_cell.number_format = Config.Export.NUMBER_FORMAT
        if 'fecha_expira' in df.columns:
            col_idx = df.columns.get_loc('fecha_expira') + 1
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for excel_cell in cell:
                    excel_cell.number_format = 'DD-MM-YYYY'
    output.seek(0)
    
    timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M")
    filename = f"inventario_stock_{timestamp}.xlsx"
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

@app.route('/dashboard', methods=['GET', 'POST'])
@require_auth
@require_admin(permission='view_dashboard')
def dashboard() -> Union[Response, str]:
    """
    Renderiza el dashboard principal con métricas de inventario y filtros.
    
    Solo accesible para usuarios con permiso 'view_dashboard'.
    Muestra resúmenes de inventario por ubicación, categoría y línea comercial.
    
    Methods:
        GET: Muestra el dashboard con filtros actuales
        POST: Aplica nuevos filtros y recarga el dashboard
    
    Form Parameters (POST):
        category_id (int, optional): ID de categoría para filtrar
        line_id (int, optional): ID de línea comercial para filtrar
        location_id (int, optional): ID de ubicación para filtrar
    
    Returns:
        str: Template HTML del dashboard con datos filtrados
        Response: Redirección si no tiene permisos
        
    Requires:
        - Sesión activa
        - Permiso: view_dashboard
    """
    # Verificar permisos de acceso al dashboard (manejado por decorador @require_admin)

    if request.method == 'POST':
        try:
            f = DashboardFilters.model_validate({
                'category_id': request.form.get('category_id'),
                'line_id': request.form.get('line_id'),
                'location_id': request.form.get('location_id'),
            })
        except ValidationError:
            abort(400)
        return redirect(url_for('dashboard',
            category_id=f.category_id,
            line_id=f.line_id,
            location_id=f.location_id,
        ))

    try:
        f = DashboardFilters.model_validate({
            'category_id': request.args.get('category_id'),
            'line_id': request.args.get('line_id'),
            'location_id': request.args.get('location_id'),
        })
    except ValidationError:
        abort(400)
    selected_category_id = f.category_id
    selected_line_id = f.line_id
    selected_location_id = f.location_id

    import time
    start_time = time.time()
    dashboard_data = data_manager.get_dashboard_data(
        category_id=selected_category_id,
        line_id=selected_line_id,
        location_id=selected_location_id
    )
    elapsed = time.time() - start_time
    
    # **OPTIMIZACIÓN**: Usamos la función get_filter_options para ser más eficientes
    filter_options = data_manager.get_filter_options()
    available_categories = filter_options.get('grupos', [])
    available_lineas = filter_options.get('lineas', [])
    available_lugares = filter_options.get('lugares', [])

    if not dashboard_data:
        flash('No hay datos de inventario para mostrar en el dashboard.', 'warning')
        return redirect(url_for('inventory'))
    
    return render_template('dashboard.html', 
        data=dashboard_data, 
        categories=available_categories, 
        lineas=available_lineas,
        lugares=available_lugares,
        selected_category_id=selected_category_id,
        selected_line_id=selected_line_id,
        selected_location_id=selected_location_id,
        backend_time=f"{elapsed:.3f} s"
    )

@app.route('/exportacion', methods=['GET', 'POST'])
@require_auth
def inventory_export() -> Union[Response, str]:
    """
    Página de inventario de la ubicación de exportación (PCP/Exportacion).
    
    Muestra productos disponibles en la ubicación específica de exportación
    con opciones de filtrado por grupo y línea comercial.
    
    Methods:
        GET: Muestra la página con filtros actuales
        POST: Aplica nuevos filtros
    
    Form Parameters (POST):
        category_id (int, optional): ID del grupo de artículos
        line_id (int, optional): ID de la línea comercial
    
    Returns:
        str: Template HTML de la página de exportación
        
    Requires:
        Sesión de usuario activa
    """
    filter_options = data_manager.get_filter_options()
    
    selected_filters = {
        'search_term': request.values.get('search_term'),
        'category_id': request.values.get('category_id', type=int),
        'line_id': request.values.get('line_id', type=int)
    }

    stock_data = data_manager.get_export_inventory(**selected_filters)
    
    return render_template(
        'export_inventory.html',
        inventory=stock_data, 
        filter_options=filter_options,
        selected_filters=selected_filters
    )

@app.route('/')
def index():
    """
    Página de inicio que redirige según el estado de la sesión.
    
    Comportamiento:
    - Si hay sesión activa: redirige al dashboard
    - Si no hay sesión: redirige al login
    
    Returns:
        Response: Redirección a dashboard o login
    """
    # Mostrar el dashboard como página principal si hay sesión, sino ir a login
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/analytics')
@require_auth
@require_admin(permission='view_analytics')
def analytics() -> Union[Response, str]:
    """
    Página de analíticas con métricas de uso de la aplicación.
    
    Solo accesible para usuarios con permiso 'view_analytics'. Muestra:
    - Total de visitas en el período seleccionado (por defecto 30 días)
    - Usuarios únicos activos
    - Total de usuarios registrados en el sistema
    - Visitas por usuario
    - Visitas por página
    - Visitas por día
    - Detalles completos de cada visita
    
    Query Parameters:
        period (int, optional): Período en días para consultar (default: 30)
    
    Returns:
        str: Template HTML con datos de analytics
        Response: Redirección si no tiene permisos
        
    Requires:
        - Sesión activa
        - Permiso: view_analytics
    """
    # Verificar permisos de administrador (manejado por decorador @require_admin)
    
    # Obtener período solicitado
    try:
        f = AnalyticsFilters.model_validate({'period': request.args.get('period', 30)})
    except ValidationError:
        abort(400)
    days = f.period
    
    # Recopilar todas las estadísticas
    total_visits = analytics_db.get_total_visits(days)
    unique_users = analytics_db.get_unique_users(days)
    
    # Obtener total de usuarios registrados desde la base de datos
    total_allowed_users = permissions_manager.get_stats()['total_active']
    
    stats = {
        'total_visits': total_visits,
        'unique_users': unique_users,
        'total_allowed_users': total_allowed_users,
        'visits_by_user': analytics_db.get_visits_by_user(days),
        'visits_by_page': analytics_db.get_visits_by_page(days),
        'visits_by_day': analytics_db.get_visits_by_day(days),
        'visits_by_hour': analytics_db.get_visits_by_hour(min(days, 7)),
        'recent_visits': analytics_db.get_recent_visits(50)
    }
    
    return render_template('analytics.html', stats=stats, period=days)

@app.route('/inventory', methods=['GET', 'POST'])
@require_auth
def inventory() -> Union[Response, str]:
    """
    Página principal de inventario con filtros múltiples y gestión de productos.
    
    Muestra el inventario completo de stock con capacidad de filtrado por:
    - Búsqueda de texto libre
    - Producto específico
    - Categoría (grupo de artículos)
    - Línea comercial
    - Ubicación de almacén
    - Estado de expiración
    
    Methods:
        GET: Muestra la página con filtros actuales
        POST: Aplica nuevos filtros
    
    Form Parameters (POST):
        category_id (int, optional): ID del grupo de artículos
        line_id (int, optional): ID de la línea comercial
        location_id (int, optional): ID de la ubicación
    
    Returns:
        str: Template HTML de la página de inventario
        Response: Redirección al login si no autenticado
        
    Requires:
        Sesión de usuario activa
    """
    # Verificar si el usuario tiene permiso de descarga
    download_whitelist = os.getenv('DOWNLOAD_WHITELIST', '').split(',') if os.getenv('DOWNLOAD_WHITELIST') else []
    can_download = session.get('username', '').lower() in [email.strip().lower() for email in download_whitelist]
    
    filter_options = data_manager.get_filter_options()
    
    src = request.form if request.method == 'POST' else request.args
    try:
        f = InventoryFilters.model_validate({
            'search_term': src.get('search_term'),
            'product_id': src.get('product_id'),
            'category_id': src.get('category_id'),
            'line_id': src.get('line_id'),
            'location_id': src.get('location_id'),
            'exp_status': request.args.get('exp_status'),
        })
    except ValidationError:
        abort(400)
    stock_data = data_manager.get_stock_inventory(
        search_term=f.search_term,
        product_id=f.product_id,
        category_id=f.category_id,
        line_id=f.line_id,
        location_id=f.location_id,
    )

    exp_status = f.exp_status
    if exp_status and stock_data:
        if exp_status == '0-3':
            stock_data = [item for item in stock_data if item['meses_expira'] is not None and 0 <= item['meses_expira'] <= 3]
        elif exp_status == '3-6':
            stock_data = [item for item in stock_data if item['meses_expira'] is not None and 3 < item['meses_expira'] <= 6]
        elif exp_status == '6-9':
            stock_data = [item for item in stock_data if item['meses_expira'] is not None and 6 < item['meses_expira'] <= 9]
        elif exp_status == '9-12':
            stock_data = [item for item in stock_data if item['meses_expira'] is not None and 9 < item['meses_expira'] <= 12]
        elif exp_status == '>12':
            stock_data = [item for item in stock_data if item['meses_expira'] is not None and item['meses_expira'] > 12]
    
    selected_filters = {
        'search_term': f.search_term,
        'product_id': f.product_id,
        'category_id': f.category_id,
        'line_id': f.line_id,
        'location_id': f.location_id,
        'exp_status': f.exp_status,
    }

    return render_template(
        'inventory.html', 
        inventory=stock_data, 
        filter_options=filter_options,
        selected_filters=selected_filters,
        can_download=can_download
    )


# =====================================================
# MÓDULO DE ADMINISTRACIÓN DE PERMISOS
# =====================================================

@app.route('/admin/users')
@require_admin('manage_users')
def admin_users() -> str:
    """
    Lista todos los usuarios del sistema con sus roles y permisos.
    
    Muestra:
    - Email del usuario
    - Rol asignado (admin_full, dashboard_user, inventory_user, viewer)
    - Estado (activo/inactivo)
    - Fecha de creación
    - Último login
    - Acciones: Editar, Eliminar, Reactivar
    
    Returns:
        str: Template HTML con la lista de usuarios
        
    Requires:
        Permiso: manage_users (solo admin_full)
    """
    # Filtros opcionales
    role_filter = request.args.get('role')
    include_inactive = request.args.get('include_inactive') == 'true'
    
    # Obtener lista de usuarios
    users = permissions_manager.list_users(
        role_filter=role_filter,
        include_inactive=include_inactive
    )
    
    # Estadísticas
    stats = permissions_manager.get_stats()
    
    # Obtener cambios recientes (última semana)
    audit_stats = audit_logger.get_stats(days_back=7)
    recent_changes = audit_stats.get('total_actions', 0)
    
    # Roles disponibles
    available_roles = list(permissions_manager.ROLE_PERMISSIONS.keys())
    
    return render_template(
        'admin/users_list.html',
        users=users,
        stats=stats,
        recent_changes=recent_changes,
        available_roles=available_roles,
        current_role_filter=role_filter,
        include_inactive=include_inactive
    )


@app.route('/admin/users/add', methods=['GET', 'POST'])
@require_admin('manage_users')
def admin_add_user() -> Union[str, Response]:
    """
    Agrega un nuevo usuario al sistema.
    
    GET: Muestra formulario de creación
    POST: Procesa la creación del usuario
    
    Form Parameters (POST):
        email (str): Email corporativo (@agrovetmarket.com)
        role (str): Rol a asignar (admin_full, dashboard_user, etc.)
    
    Returns:
        str: Template HTML del formulario
        Response: Redirección a lista de usuarios si se creó exitosamente
        
    Requires:
        Permiso: manage_users (solo admin_full)
    """
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        role = request.form.get('role', '').strip()
        
        # Validaciones
        if not email or not role:
            flash('Email y rol son requeridos', 'error')
            return redirect(url_for('admin_add_user'))
        
        if not email.endswith('@agrovetmarket.com'):
            flash('Solo se permiten emails corporativos @agrovetmarket.com', 'error')
            return redirect(url_for('admin_add_user'))
        
        if role not in permissions_manager.ROLE_PERMISSIONS:
            flash(f'Rol inválido: {role}', 'error')
            return redirect(url_for('admin_add_user'))
        
        # Crear usuario
        created_by = session.get('username')
        try:
            success = permissions_manager.add_user(email, role, created_by)
            
            if success:
                # Registrar en audit log
                audit_logger.log_user_created(
                    user_email=email,
                    role=role,
                    created_by=created_by,
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string
                )
                
                flash(f'Usuario {email} creado exitosamente con rol {role}', 'success')
                logger.info(f"[ADMIN] Usuario creado: {email} ({role}) por {created_by}")
            else:
                flash(f'El usuario {email} ya existe', 'warning')
            
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            logger.error(f"[ADMIN] Error al crear usuario {email}: {e}")
            flash(f'Error al crear usuario: {str(e)}', 'error')
            return redirect(url_for('admin_add_user'))
    
    # GET: Mostrar formulario
    available_roles = list(permissions_manager.ROLE_PERMISSIONS.keys())
    role_permissions = permissions_manager.ROLE_PERMISSIONS
    
    return render_template(
        'admin/user_add.html',
        available_roles=available_roles,
        role_permissions=role_permissions
    )


@app.route('/admin/users/edit/<email>', methods=['GET', 'POST'])
@require_admin('manage_users')
def admin_edit_user(email: str) -> Union[str, Response]:
    """
    Edita el rol de un usuario existente.
    
    GET: Muestra formulario de edición
    POST: Procesa la actualización del rol
    
    Args:
        email: Email del usuario a editar
    
    Form Parameters (POST):
        new_role (str): Nuevo rol a asignar
    
    Returns:
        str: Template HTML del formulario
        Response: Redirección a lista de usuarios si se actualizó exitosamente
        
    Requires:
        Permiso: manage_users (solo admin_full)
        
    Security:
        - No permite auto-modificación (previene escalada de privilegios)
        - Registra todos los cambios en audit log
    """
    email = email.lower()
    current_user = session.get('username')
    
    # Prevenir auto-modificación
    if email == current_user:
        flash('No puedes modificar tu propio rol', 'error')
        return redirect(url_for('admin_users'))
    
    # Obtener datos del usuario
    user = permissions_manager.get_user_details(email)
    if not user:
        flash(f'Usuario {email} no encontrado', 'error')
        return redirect(url_for('admin_users'))
    
    if request.method == 'POST':
        new_role = request.form.get('new_role', '').strip()
        
        if new_role not in permissions_manager.ROLE_PERMISSIONS:
            flash(f'Rol inválido: {new_role}', 'error')
            return redirect(url_for('admin_edit_user', email=email))
        
        old_role = user['role']
        
        if old_role == new_role:
            flash('El rol no ha cambiado', 'info')
            return redirect(url_for('admin_users'))
        
        try:
            success = permissions_manager.update_user_role(email, new_role, current_user)
            
            if success:
                # Registrar en audit log
                audit_logger.log_user_updated(
                    user_email=email,
                    old_role=old_role,
                    new_role=new_role,
                    updated_by=current_user,
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string
                )
                
                flash(f'Rol de {email} actualizado: {old_role} → {new_role}', 'success')
                logger.info(f"[ADMIN] Rol actualizado: {email} {old_role}→{new_role} por {current_user}")
            else:
                flash(f'Error al actualizar rol de {email}', 'error')
            
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            logger.error(f"[ADMIN] Error al actualizar rol de {email}: {e}")
            flash(f'Error al actualizar rol: {str(e)}', 'error')
            return redirect(url_for('admin_edit_user', email=email))
    
    # GET: Mostrar formulario
    available_roles = list(permissions_manager.ROLE_PERMISSIONS.keys())
    role_permissions = permissions_manager.ROLE_PERMISSIONS
    
    return render_template(
        'admin/user_edit.html',
        user=user,
        available_roles=available_roles,
        role_permissions=role_permissions
    )


@app.route('/admin/users/delete/<email>', methods=['POST'])
@require_admin('manage_users')
def admin_delete_user(email: str) -> Response:
    """
    Desactiva un usuario (soft delete).
    
    Args:
        email: Email del usuario a desactivar
    
    Returns:
        Response: Redirección a lista de usuarios
        
    Requires:
        Permiso: manage_users (solo admin_full)
        
    Security:
        - No permite auto-eliminación
        - Soft delete (is_active=False) para mantener historial
        - Registra la acción en audit log
    """
    email = email.lower()
    current_user = session.get('username')
    
    # Prevenir auto-eliminación
    if email == current_user:
        flash('No puedes desactivar tu propia cuenta', 'error')
        return redirect(url_for('admin_users'))
    
    # Obtener datos del usuario antes de eliminar
    user = permissions_manager.get_user_details(email)
    if not user:
        flash(f'Usuario {email} no encontrado', 'error')
        return redirect(url_for('admin_users'))
    
    try:
        success = permissions_manager.delete_user(email, current_user)
        
        if success:
            # Registrar en audit log
            audit_logger.log_user_deleted(
                user_email=email,
                role=user['role'],
                deleted_by=current_user,
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string
            )
            
            flash(f'Usuario {email} desactivado exitosamente', 'success')
            logger.info(f"[ADMIN] Usuario desactivado: {email} por {current_user}")
        else:
            flash(f'Error al desactivar usuario {email}', 'error')
        
    except Exception as e:
        logger.error(f"[ADMIN] Error al desactivar usuario {email}: {e}")
        flash(f'Error al desactivar usuario: {str(e)}', 'error')
    
    return redirect(url_for('admin_users'))


@app.route('/admin/users/reactivate/<email>', methods=['POST'])
@require_admin('manage_users')
def admin_reactivate_user(email: str) -> Response:
    """
    Reactiva un usuario desactivado.
    
    Args:
        email: Email del usuario a reactivar
    
    Returns:
        Response: Redirección a lista de usuarios
        
    Requires:
        Permiso: manage_users (solo admin_full)
    """
    email = email.lower()
    current_user = session.get('username')
    
    # Obtener datos del usuario
    user = permissions_manager.get_user_details(email)
    if not user:
        flash(f'Usuario {email} no encontrado', 'error')
        return redirect(url_for('admin_users'))
    
    try:
        success = permissions_manager.reactivate_user(email, current_user)
        
        if success:
            # Registrar en audit log
            audit_logger.log_user_reactivated(
                user_email=email,
                role=user['role'],
                reactivated_by=current_user,
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string
            )
            
            flash(f'Usuario {email} reactivado exitosamente', 'success')
            logger.info(f"[ADMIN] Usuario reactivado: {email} por {current_user}")
        else:
            flash(f'Error al reactivar usuario {email}', 'error')
        
    except Exception as e:
        logger.error(f"[ADMIN] Error al reactivar usuario {email}: {e}")
        flash(f'Error al reactivar usuario: {str(e)}', 'error')
    
    return redirect(url_for('admin_users'))


@app.route('/admin/audit-log')
@require_admin('view_audit_log')
def admin_audit_log() -> str:
    """
    Muestra el historial de auditoría de cambios en permisos.
    
    Query Parameters:
        action (str, optional): Filtrar por tipo de acción
        user_email (str, optional): Filtrar por email de usuario
        days (int, optional): Días hacia atrás (default: 30)
        limit (int, optional): Número máximo de registros (default: 100)
    
    Returns:
        str: Template HTML con el historial de auditoría
        
    Requires:
        Permiso: view_audit_log (solo admin_full)
    """
    # Parámetros de filtrado
    action_filter = request.args.get('action')
    user_email_filter = request.args.get('user_email')
    days_back = int(request.args.get('days', 30))
    limit = int(request.args.get('limit', 100))
    
    # Obtener logs
    logs = audit_logger.get_logs(
        limit=limit,
        action_filter=action_filter,
        user_email_filter=user_email_filter,
        days_back=days_back
    )
    
    # Estadísticas
    stats = audit_logger.get_stats(days_back=7)  # Última semana
    
    # Tipos de acciones disponibles para filtros
    action_types = ['user_created', 'role_changed', 'user_deleted', 'user_reactivated', 'login']
    
    return render_template(
        'admin/audit_log.html',
        logs=logs,
        stats=stats,
        action_types=action_types,
        current_action_filter=action_filter,
        current_user_filter=user_email_filter,
        days_back=days_back
    )


if __name__ == '__main__':
    # Modo debug condicional basado en variable de entorno (OWASP A05: Security Misconfiguration)
    debug_mode = os.getenv('ENVIRONMENT', 'development') == 'development'
    host = '0.0.0.0' if os.getenv('ENVIRONMENT') == 'production' else '127.0.0.1'  # nosec B104 — intencional en producción bajo WSGI/reverse proxy
    
    if debug_mode:
        logger.info("🛠️ MODO DESARROLLO: Debug activado")
    else:
        logger.info("🔒 MODO PRODUCCIÓN: Debug desactivado")
    
    app.run(debug=debug_mode, host=host, port=5000)